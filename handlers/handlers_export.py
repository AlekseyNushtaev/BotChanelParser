from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import asyncio

from config import ADMIN_IDS
from logger import logger
from db.posts import get_posts

export_router = Router()


async def create_excel_file() -> str:
    """
    Создает Excel файл со всеми записями из таблицы Post
    """
    # Получаем все посты из базы данных
    posts = await get_posts(limit=10000)

    # Создаем новую рабочую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts Export"

    # Заголовки столбцов
    headers = [
        "ID", "Chat ID", "Chat Title", "Chat Type", "Message ID",
        "Content Type", "Text", "Telegram File IDs", "Digest", "AI Generated",
        "Original Date", "Received At", "Processed At"
    ]

    # Настраиваем стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Записываем данные
    for row_num, post in enumerate(posts, 2):
        # Преобразуем JSON поля в строки
        telegram_file_ids = str(post.telegram_file_ids) if post.telegram_file_ids else ""

        # Преобразуем даты в строки
        original_date = post.original_date.strftime("%Y-%m-%d %H:%M:%S") if post.original_date else ""
        received_at = post.received_at.strftime("%Y-%m-%d %H:%M:%S") if post.received_at else ""
        processed_at = post.processed_at.strftime("%Y-%m-%d %H:%M:%S") if post.processed_at else ""

        # Очищаем HTML теги из текста для Excel
        text = post.text or ""
        if text:
            import re
            text = re.sub(r'<[^>]+>', '', text)  # Удаляем HTML теги
            text = re.sub(r'&[a-z]+;', '', text)  # Удаляем HTML entities

        # Заполняем строку
        ws.cell(row=row_num, column=1, value=post.id)
        ws.cell(row=row_num, column=2, value=post.chat_id)
        ws.cell(row=row_num, column=3, value=post.chat_title)
        ws.cell(row=row_num, column=4, value=post.chat_type)
        ws.cell(row=row_num, column=5, value=post.message_id)
        ws.cell(row=row_num, column=6, value=post.content_type)
        ws.cell(row=row_num, column=7, value=text)  # Используем очищенный текст
        ws.cell(row=row_num, column=8, value=telegram_file_ids)
        ws.cell(row=row_num, column=9, value="Да" if post.digest else "Нет")
        ws.cell(row=row_num, column=10, value=post.ai_gen)
        ws.cell(row=row_num, column=11, value=original_date)
        ws.cell(row=row_num, column=12, value=received_at)
        ws.cell(row=row_num, column=13, value=processed_at)

    # Настраиваем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Добавляем фильтры к заголовкам
    ws.auto_filter.ref = ws.dimensions

    # Создаем имя файла с временной меткой
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"posts_export_{timestamp}.xlsx"
    filepath = f"temp/{filename}"

    # Создаем временную директорию, если её нет
    os.makedirs("temp", exist_ok=True)

    # Сохраняем файл
    wb.save(filepath)

    return filepath


@export_router.message(Command("export_posts"))
async def export_posts_command(message: Message, state: FSMContext):
    """
    Команда для экспорта всех постов в Excel файл
    """
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("Доступ запрещен")
        return

    try:
        await state.clear()

        # Отправляем сообщение о начале процесса
        processing_msg = await message.answer("⏳ Начинаю экспорт данных... Это может занять некоторое время.")

        # Создаем Excel файл
        filepath = await create_excel_file()

        # Отправляем файл пользователю
        file = FSInputFile(filepath, filename=f"posts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        await message.answer_document(
            document=file,
            caption=f"✅ Экспорт завершен!\n"
                    f"📊 Всего записей: {len(await get_posts(limit=10000))}\n"
                    f"📅 Дата экспорта: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

        # Удаляем сообщение о процессе
        await processing_msg.delete()

        logger.info(f"[{message.from_user.id}] Экспорт постов выполнен успешно. Файл: {filepath}")

        # Удаляем временный файл через некоторое время
        async def cleanup_file():
            await asyncio.sleep(10)
            if os.path.exists(filepath):
                os.remove(filepath)
                logger.info(f"Временный файл удален: {filepath}")

        asyncio.create_task(cleanup_file())

    except Exception as e:
        logger.error(f"[{message.from_user.id}] Ошибка при экспорте постов: {e}")
        await message.answer(f"❌ Ошибка при экспорте данных: {str(e)}")


@export_router.message(Command("stats"))
async def show_stats_command(message: Message, state: FSMContext):
    """
    Команда для показа статистики по постам
    """
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("Доступ запрещен")
        return

    try:
        await state.clear()

        # Получаем все посты
        posts = await get_posts(limit=10000)

        if not posts:
            await message.answer("📭 В базе данных нет записей")
            return

        # Подсчитываем статистику
        total_posts = len(posts)

        # Статистика по типам контента
        content_types = {}
        for post in posts:
            content_type = post.content_type or "unknown"
            content_types[content_type] = content_types.get(content_type, 0) + 1

        # Статистика по чатам
        chats = {}
        for post in posts:
            chat_title = post.chat_title or f"ID: {post.chat_id}"
            chats[chat_title] = chats.get(chat_title, 0) + 1

        # Количество постов в дайджесте
        in_digest = sum(1 for post in posts if post.digest)

        # Формируем сообщение
        stats_message = (
            f"📊 Статистика по постам:\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📂 Всего постов: {total_posts}\n"
            f"📋 В дайджесте: {in_digest}\n\n"
            f"📈 Типы контента:\n"
        )

        for content_type, count in sorted(content_types.items(), key=lambda x: x[1], reverse=True)[:5]:
            percentage = (count / total_posts) * 100
            stats_message += f"  • {content_type}: {count} ({percentage:.1f}%)\n"

        stats_message += f"\n📁 Топ чатов:\n"
        for chat_title, count in sorted(chats.items(), key=lambda x: x[1], reverse=True)[:5]:
            percentage = (count / total_posts) * 100
            stats_message += f"  • {chat_title}: {count} ({percentage:.1f}%)\n"

        stats_message += f"\n💾 Для полного экспорта используйте /export_posts"

        await message.answer(stats_message)

        logger.info(f"[{message.from_user.id}] Статистика показана успешно")

    except Exception as e:
        logger.error(f"[{message.from_user.id}] Ошибка при получении статистики: {e}")
        await message.answer(f"❌ Ошибка при получении статистики: {str(e)}")